import cv2
import dlib
import time
import threading
import math
import datetime
import os
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from skimage import util
from skimage.io import imread
from skimage.filters import threshold_otsu
from skimage import measure
from skimage.measure import regionprops
import matplotlib.patches as patches
import sklearn
import joblib
#import trial_23


SNo = int(1)
counter = 0
sheet = {}
now = datetime.datetime.now()

mainFilename = 'C:\\Users\SHREYA\Desktop\mainFileVehicleID.xlsx'

if not(os.path.isfile(mainFilename)):
    mwb = openpyxl.Workbook() #main work book
    #main_sheet = mwb.create_sheet('main_sheet')
    #main_sheet = mwb.worksheets
    main_sheet=mwb.active  
    #main_sheet = main_sheet.title
    main_sheet.cell(row=1, column=1).value = 'VehicleID'
    main_sheet.cell(row=1, column=2).value = 'Vehicle_Number'


    sheet[0] = mwb.create_sheet('sheet%d' %counter)
    # ws[counter] = mwb.worksheets[counter]
    sheet[0] = mwb.active
    a = 'Daily Vehicles'
    day = now.day
    b = str(str(a) + str(day) + '.' + str(now.month) + '.' + str(now.year))
    sheet[0].title = b
    sheet[0].cell(row=1, column=1).value = 'VehicleID'
    sheet[0].cell(row=1, column=2).value = 'Date'
    sheet[0].cell(row=1, column=3).value = 'Time'
    sheet[0].cell(row=1, column=4).value = 'Camera'
    sheet[0].cell(row=1, column=5).value = 'Speed'
    sheet[0].cell(row=1, column=6).value = 'Number Plate'

    mwb.save(mainFilename)

carCascade = cv2.CascadeClassifier('C:\\Users\\SHREYA\\Desktop\\myhaar.xml')
video = cv2.VideoCapture('C:\\Users\\SHREYA\\Desktop\\shreya\\Internship\\Videos\\a_1.avi')

WIDTH = 1280
HEIGHT = 720

######

def detection(image_1):
    #print(img)
    height = 900
    width = 700
    img = cv2.cvtColor(image_1, cv2.COLOR_RGB2GRAY)
    car_image = cv2.resize(img, (width,height))
    img_1 = cv2.resize(img, (width,height))
    # car_image = imread("car.png", as_gray=True)
    # it should be a 2 dimensional array

    # the next line is not compulsory however, a grey scale pixel
    # in skimage ranges between 0 & 1. multiplying it with 255
    # will make it range between 0 & 255 (something we can relate better with
    


    gray_car_image = car_image * 255
    #fig, (ax1, ax2) = plt.subplots(1, 2)
    #ax1.imshow(gray_car_image, cmap="gray")
    threshold_value = threshold_otsu(gray_car_image)
    binary_car_image = gray_car_image > threshold_value
    #print(binary_car_image)
    #ax2.imshow(binary_car_image, cmap="gray")
    # ax2.imshow(gray_car_image, cmap="gray")
    #plt.show()

    # CCA (finding connected regions) of binary image


    # this gets all the connected regions and groups them together
    label_image = measure.label(binary_car_image)

    #print(label_image.shape[0]) #height of the image
    #print(label_image.shape[1]) #width of the image
    # getting the maximum width, height and minimum width and height that a license plate can be

    #plate_dimensions0 = (0.01*label_image.shape[0], 0.03*label_image.shape[0], 0.10*label_image.shape[1], 0.15*label_image.shape[1])
    plate_dimensions = (0.06*label_image.shape[0], 0.13*label_image.shape[0], 0.085*label_image.shape[1], 0.2*label_image.shape[1])
    plate_dimensions2 = (0.06*label_image.shape[0], 0.18*label_image.shape[0], 0.15*label_image.shape[1], 0.25*label_image.shape[1])
    min_height, max_height, min_width, max_width = plate_dimensions
    plate_objects_cordinates = []
    plate_like_objects = []

    #print(min_height, ' ', max_height, ' ', min_width, ' ', max_width)

    #fig, (ax1) = plt.subplots(1)
    #ax1.imshow(gray_car_image, cmap="gray")
    flag =0
    go = 0
    # regionprops creates a list of properties of all the labelled regions

    for region in regionprops(label_image):
    
        #min_row, min_col, max_row, max_col = region.bbox
        #print('')
        #print('min_row:', min_row, 'max_row:', max_row, 'min_col:', min_col, 'max_col:', max_col)
        #print(region.area)
        #print('')
        
        if (region.area < 500):
            continue
            # 50 -> 30
            #print(region.area)
            #min_row, min_col, max_row, max_col = region.bbox
            #region_height = max_row - min_row
            #region_width = max_col - min_col
        
            #if the region is so small then it's likely not a license plate
            # the bounding box coordinates
        print(type(region.bbox))
        print(region.bbox)
        
        min_row, min_col, max_row, max_col = region.bbox
        # print(min_row)
        # print(min_col)
        # print(max_row)
        # print(max_col)
        #print(yo+1)
    
        region_height = max_row - min_row
        region_width = max_col - min_col
        #print('')
        #print('region_height: ', region_height, 'region_width: ', region_width)
        #print('min_row:', min_row, 'max_row:', max_row, 'min_col:', min_col, 'max_col:', max_col)
        #print('region_width: ', region_width)
        #print('')

        # ensuring that the region identified satisfies the condition of a typical license plate
        if (region_height >= min_height and region_height <= max_height and region_width >= min_width and region_width <= max_width and region_width > region_height):
            flag = 1
            go = go + 1
            print(go)
        
            #fig, (ax1) = plt.subplots(1)
            #ax1.imshow(gray_car_image, cmap="gray")
        
            plate_like_objects.append(binary_car_image[min_row:max_row, min_col:max_col])
            plate_objects_cordinates.append((min_row, min_col, max_row, max_col))
            rectBorder = patches.Rectangle((min_col, min_row), max_col - min_col, max_row - min_row, edgecolor="red", linewidth=2, fill=False)
            #ax1.add_patch(rectBorder)
            #plt.show()
            
            #cropped_image = gray_car_image[min_row:max_row, min_col:max_col]
            #print(type(cropped_image))
            #changed_color = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
    
            #inverted_image = util.invert(cropped_image)
            #threshold_value_license_plate = threshold_otsu(inverted_image)
            #print(threshold_value_license_plate)
            #thresholded_image = inverted_image > threshold_value_license_plate
            
            #fig, (ax1) = plt.subplots(1)
            #ax1.imshow(thresholded_image, cmap='gray')
            #plt.show()
            
            # let's draw a red rectangle over those regions
    
    if(flag == 1):
        print(plate_like_objects[0])
        #plt.show()
        #print('flag=1')


    if(flag == 0):
        min_height, max_height, min_width, max_width = plate_dimensions2
        plate_objects_cordinates = []
        plate_like_objects = []

        #fig, (ax1) = plt.subplots(1)
        #ax1.imshow(gray_car_image, cmap="gray")

        # regionprops creates a list of properties of all the labelled regions
        for region in regionprops(label_image):
            if (region.area < 150): ## 50 -> 30
                #if the region is so small then it's likely not a license plate
                continue
                # the bounding box coordinates
            min_row, min_col, max_row, max_col = region.bbox
            # print(min_row)
            # print(min_col)
            # print(max_row)
            # print(max_col)
    
            region_height = max_row - min_row
            region_width = max_col - min_col
            # print(region_height)
            # print(region_width)
    
            # ensuring that the region identified satisfies the condition of a typical license plate
            if (region_height >= min_height and region_height <= max_height and region_width >= min_width and region_width <= max_width and region_width > region_height):
                # print("hello")
                plate_like_objects.append(binary_car_image[min_row:max_row, min_col:max_col])
                plate_objects_cordinates.append((min_row, min_col, max_row, max_col))
                rectBorder = patches.Rectangle((min_col, min_row), max_col - min_col, max_row - min_row, edgecolor="red", linewidth=2, fill=False)
                #ax1.add_patch(rectBorder)
                #plt.show()
                #print('flag=0')
                
                #cropped_image = gray_car_image[min_row:max_row, min_col:max_col]
                #print(type(cropped_image))
                #changed_color = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
        
                #inverted_image = util.invert(cropped_image)
                #fig, (ax1) = plt.subplots(1)
                #ax1.imshow(inverted_image, cmap='gray')
                #plt.show()
                # let's draw a red rectangle over those regions
    
        #print(plate_like_objects[3])
        #plt.show()
    
    #print(flag)
    #print(len(plate_like_objects))
    #print(plate_like_objects[len(plate_like_objects)-1])
    #print(plate_objects_cordinates[len(plate_like_objects)-1])

    if (len(plate_like_objects)!=0):
        print(len(plate_like_objects))
    
        (r1, c1, r2, c2)=plate_objects_cordinates[len(plate_like_objects)-1]
    
        cropped_image_1 = gray_car_image[r1:r2, c1:c2]
        img_2 = img_1[r1:r2, c1:c2]
        img_3 = cv2.medianBlur(img_2,5)
    #print(cropped_image)
    #grayscale = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
        ret,th_1 = cv2.threshold(img_3,155,255,cv2.THRESH_BINARY)
    
        th_2 = cv2.adaptiveThreshold(img_3,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,11,2)# this one is better after all the thresholdings
    
    #th_3 = cv2.adaptiveThreshold(img_3,255,cv2.ADAPTIVE_THRESH_MEAN_C,\
    #            cv2.THRESH_BINARY,11,2)
    
        not_th_2 = cv2.bitwise_not(th_2)
    
    #fig, (ax1,ax2) = plt.subplots(1,2)
    #ax1.imshow(th_2, cmap='gray')
    #ax2.imshow(not_th_2, cmap='gray')
    #plt.show()
    
        width_plate = 90
        height_plate = 60
    
        license_plate = cv2.resize(not_th_2, (width_plate,height_plate))
        labelled_plate = measure.label(license_plate)
    #print(license_plate.shape[0]) #height of the plate
        
        character_dimensions = (0.1875*license_plate.shape[0], 0.480*license_plate.shape[0], 0.05*license_plate.shape[1], 0.40*license_plate.shape[1])
        min_height, max_height, min_width, max_width = character_dimensions #from resized license plate
    #print(character_dimensions)
    
        characters = []
        counter=0
        column_list = []
    
    #fig, (ax1) = plt.subplots(1)
    #ax1.imshow(license_plate, cmap="gray")
        
        for regions in regionprops(labelled_plate):
            y0, x0, y1, x1 = regions.bbox
            region_height = y1 - y0
            region_width = x1 - x0

    #print(region.bbox)
        #print('')
        #print(region.bbox)
        #print('')
        #print('region_height', region_height, 'region_width', region_width)
        #print('min_height', min_height, 'max_height', max_height, 'min_width', min_width, 'max_width', max_width)
        
            if (region_height > min_height and region_height < max_height and region_width > min_width and region_width < max_width):
                roi = license_plate[y0:y1, x0:x1]
    
            #print('')
            #print(roi)
            #print('')
            
            # draw a red bordered rectangle over the character.
                rect_border = patches.Rectangle((x0, y0), x1 - x0, y1 - y0, edgecolor="red", linewidth=2, fill=False)
            #print(rect_border)
            
            #ax1.add_patch(rect_border)
    
            # resize the characters to 20X20 and then append each character into the characters list
                resized_char = cv2.resize(roi, (20, 20))
                characters.append(resized_char)
            #print('')
            #print(resized_char)
            
            # this is just to keep track of the arrangement of the characters
                column_list.append(x0)
    
            #print(column_list)
    # print(characters)
    #plt.show()
    
    #print("Loading model")
    #filename = 'C:\\Users\\SHREYA\\Desktop\\shreya\\Internship\\finalized_model.sav' # write the full directory path and also put double slashes if required!
    #model = pickle.load(open(filename, 'rb'))
    #print('Model loaded. Predicting characters of number plate')
    
    #print("Loading model")
        filename = 'C:\\Users\\SHREYA\\Desktop\\finalized_model.sav' # write the full directory path and also put double slashes if required!
    
        current_dir = os.path.dirname(os.path.realpath(filename))
    
        model_dir = os.path.join(current_dir, 'finalized_model.sav')
    
        model = joblib.load(model_dir)
    
    #print('Model loaded. Predicting characters of number plate')
    
        classification_result = []
        
        for each_character in characters:
        # converts it to a 1D array
            each_character = each_character.reshape(1, -1);
        #print(each_character)
            result = model.predict(each_character)
            classification_result.append(result)
    
    #print('Classification result')
    #print(classification_result)
    
        plate_string = ''
    
        for eachPredict in classification_result:
            plate_string += eachPredict[0]
    
    #print('Predicted license plate')
    #print(plate_string)
        
    # it's possible the characters are wrongly arranged
    # since that's a possibility, the column_list will be
    # used to sort the letters in the right order

        column_list_copy = column_list[:]
        column_list.sort()
        rightplate_string = ''

        for each in column_list:
            rightplate_string += plate_string[column_list_copy.index(each)]

    #print('License plate')
    #print(rightplate_string)

    else:
        plate_string = 'not able to detect'

    return plate_string



######

def estimateSpeed(location1, location2, fps):
    d_pixels = math.sqrt(math.pow(location2[0] - location1[0], 2) + math.pow(location2[1] - location1[1], 2))
    # ppm = location2[2] / carWidht
    ppm = 10                          #$$$$
    d_meters = d_pixels / ppm
    #print("d_pixels=" + str(d_pixels), "d_meters=" + str(d_meters))
    #fps = 10#$$$$
    print(fps)
    speed = d_meters * 10 * 3.6 #3600/1000
    return speed
	

def trackMultipleObjects():
    rectangleColor = (0, 255, 0)
    frameCounter = 0
    currentCarID = 0
    fps = 1
    cropped_image = {}
    LPN = {}
    carTracker = {}
    carNumbers = {}
    carLocation1 = {}
    carLocation2 = {}
    speed = [None] * 1000
    counter = 0
    a = 0
    myuse = []
    	
    # Write output to video file
    out = cv2.VideoWriter('C:\\Users\\SHREYA\\Desktop\\outpy.avi',cv2.VideoWriter_fourcc('M','J','P','G'), 10, (WIDTH,HEIGHT))


    while True:

        now = datetime.datetime.now()
        if (now.hour == 0 & now.minute == 0 & now.second == 0):
            a = 1
            counter = counter + 1
            filename = 'C:\\Users\SHREYA\Desktop\mainFileVehicleID.xlsx'
            mwb = load_workbook(filename)
            sheet[counter] = mwb.create_sheet('sheet%d' %counter)
            # ws[counter] = mwb.worksheets[counter]
            sheet[counter] = mwb.active
            a = 'Daily Vehicles'
            day = now.day
            b = str(str(a) + str(day) + '.' + str(now.month) + '.' + str(now.year))
            sheet[counter].title = b
            sheet[counter].cell(row=1, column=1).value = 'VehicleID'
            sheet[counter].cell(row=1, column=2).value = 'Date'
            sheet[counter].cell(row=1, column=3).value = 'Time'
            sheet[counter].cell(row=1, column=4).value = 'Camera'
            sheet[counter].cell(row=1, column=5).value = 'Speed'
            sheet[counter].cell(row=1, column=6).value = 'Number Plate'
            mwb.save(filename)
                
        start_time = time.time()
        rc, image_2 = video.read()
        if type(image_2) == type(None):
            break
		
        image_1 = cv2.resize(image_2, (WIDTH, HEIGHT))
        resultImage = image_1.copy()
        height = image_1.shape[0]
        width = image_1.shape[1]
        image = image_1[0:height, 0:int(width/2)]
		
        frameCounter = frameCounter + 1
		
        carIDtoDelete = []

        for carID in carTracker.keys():
            trackingQuality = carTracker[carID].update(image)
			
            if trackingQuality < 7: #7
                carIDtoDelete.append(carID)
				
        for carID in carIDtoDelete:
            #print ('Removing carID ' + str(carID) + ' from list of trackers.')
            #print ('Removing carID ' + str(carID) + ' previous location.')
            #print ('Removing carID ' + str(carID) + ' current location.')
            carTracker.pop(carID, None)
            carLocation1.pop(carID, None)
            carLocation2.pop(carID, None)
            cropped_image.pop(carID,None)
            LPN.pop(carID, None)
		
        if not (frameCounter % 10): #it goes inside only when frameCounter is a multiple of 10
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            cars = carCascade.detectMultiScale(gray, 1.1, 13, 18, (24, 24))

            cars = list(cars)

            #print(cars)
            
            for i in range (len(cars)):
                #print(cars, 'HAKUNA MATATA')
                if (cars[i][0] < 195) or (cars[i][2] < 58) or (cars[i][3] < 58) or ((cars[i][1] > 40) and (cars[i][2] < 68)) or ((cars[i][1] > 268) and (cars[i][2] < 250)) or ((cars[i][1] > 268) and (cars[i][3] < 250)) :
                    myuse.append(i)
                    #print('1st', len(cars), ' ', myuse)
                    #print('SDJVFVJKS')

                #if (cars[i][2] < 58) or (cars[i][3] < 58):
                    #myuse.append(i)

                #if (cars[i][0] > 180) and (cars[i][3] < 150):
                    #myuse.append(i)

                    
            #print(cars, 'djsidSHJiiiIINHDHNJ')
            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            for i in range(len(cars)):
                if (cars[i][1] > 40) and (cars[i][2] < 70):
                    myuse.append(i)

            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            for i in range(len(cars)):
                if (cars[i][1] > 110) and (cars[i][2] < 120):
                    myuse.append(i)

            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            for i in range(len(cars)):
                if (cars[i][1] > 268) and (cars[i][2] < 250):
                    myuse.append(i)

            if myuse != []:
                #print(cars)
                for i in range(len(myuse)-1, -1, -1):
                    cars.pop(myuse[i])

                #print(cars)
                myuse = []

            #print(cars)

            #for x in range(len(cars)):
                #x_1 = cars[x][0]
                #y_1 = cars[x][1]
                #w_1 = cars[x][2]
                #h_1 = cars[x][3]
                #for y in range(len(cars)):
                    #if (x!=y):
                        #x_2 = cars[y][0]
                        #y_2 = cars[y][1]
                        #w_2 = cars[y][2]
                        #h_2 = cars[y][3]

                        #if ((x_2 > (x_1 + 5) > (x_2 + w_2)) and (x_2 > (x_1 + w_1 - 5) > (x_2 + w_2)) and (y_2 > (y_1 + h_1 + 5) > (y_2 + h_2))) :
                            #myuse.append(x)
                            #print('HEY THERE IM RIGHT HERE BONJOIRs')

            #if myuse != []:
                #print(cars)
                #for i in range(len(myuse)-1, -1, -1):
                    #cars.pop(myuse[i])

            #myuse = []         
            #for i in range(len(cars)):
                #if (cars[i][2] < 58) or (cars[i][3] < 58):
                    #myuse.append(i)
                    #print('2nd', len(cars), ' ', myuse)

            #if myuse != []:
                #for yo in myuse:
                    #if (yo < len(cars)):
                        #cars.pop(yo)

                #myuse = []

            #for i in range(len(cars)):
            
                
            #for (_x, _y, _w, _h) in cars:
                #print('yobaby')
                #print(_x, _y, _w, _h)
                #if (_h > int(height/2)):
                    #cv2.imshow()
                
                #for (try_x, try_y, try_w, try_h) in cars:
                    #if (((_x > try_x) and (_x < (try_x + try_w)) and ((_y + _h) > try_y) and ((_y + _h) < (try_y + try_h))) or ((_x > try_x) and (_x < (try_x + try_w)) and (_y > try_y) and (_y < (try_y + try_h))) or (((_x + _w) > try_x) and ((_x + _w) < try_x) and (_y > try_y) and (_y < (try_y + try_h))) or (((_x + _w) > try_x) and ((_x + _w) < (try_x + try_w)) and ((_y + _h) > try_y) and ((_y + _h) < (try_y + try_h)))):
                        #print('HEY THERE IM RIGHT HERE BONJOIRs')
                        #if ((_w * _h) < (try_w * try_h)):
                            #for i in range(len(cars)):
                                #print(cars[i][0], 'EHFHREHHSHJFGNF')
                                #print('BIENTOT MASALA')
                                #if (_x == cars[i][0]) and (_y == cars[i][1]) and (_w == cars[i][2]) and (_h == cars[i][3]):
                                    #print('oui, a FILLE')
                                #if ([_x _y _w _h] == cars[i]):
                                    #a.append(i)
                                    #print('MERA PIYA GHAR AAYA OO RAMJI')
                        #for i in range(len(cars)):
                            #if 
			
            for (_x, _y, _w, _h) in cars:
                x = int(_x)
                y = int(_y)
                w = int(_w)
                h = int(_h)
			
                x_bar = x + 0.5 * w
                y_bar = y + 0.5 * h
				
                matchCarID = None
                
		
                for carID in carTracker.keys():
                    trackedPosition = carTracker[carID].get_position()
		    
                    t_x = int(trackedPosition.left())
                    t_y = int(trackedPosition.top())
                    t_w = int(trackedPosition.width())
                    t_h = int(trackedPosition.height())

                    					
                    t_x_bar = t_x + 0.5 * t_w
                    t_y_bar = t_y + 0.5 * t_h
				
                    if ((t_x <= x_bar <= (t_x + t_w)) and (t_y <= y_bar <= (t_y + t_h)) and (x <= t_x_bar <= (x + w)) and (y <= t_y_bar <= (y + h))):
                        matchCarID = carID
				
                if matchCarID is None:
                    print ('Creating new tracker ' + str(currentCarID))
				    
                    tracker = dlib.correlation_tracker()
                    tracker.start_track(image, dlib.rectangle(x, y, x + w, y + h))
					
                    carTracker[currentCarID] = tracker
                    carLocation1[currentCarID] = [x, y, w, h]

                    currentCarID = currentCarID + 1
		
		#cv2.line(resultImage,(0,480),(1280,480),(255,0,0),5)

        myuse = []
        #print(cars)
        #print(carTracker, 'BEFOREEEE')

        ###1
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')
                    if (x_2 < (x_1 + 10)) and ((x_1 + 10) < (x_2 + w_2)) and ((x_1 + w_1 - 10) > x_2) and ((x_1 + w_1 - 10) < (x_2 + w_2)) and ((y_1 + h_1 + 10) > y_2) and ((y_1 + h_1 +10) < (y_2 + h_2)):
                        myuse.append(carID)
                        #print('HAKUNAAAAA')

        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        ###2
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')

                    if ((x_1 + w_1 + 10) > x_2) and ((x_1 + w_1 + 10) < (x_2 + w_2)) and ((y_1 + 10) > y_2) and ((y_1 + 10) < (y_2 + h_2)) and ((y_1 + h_1 - 10) > y_2) and ((y_1 + h_1 - 10) < (y_2 + h_2)) :
                        myuse.append(carID)


        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        ###3
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')

                    if ((x_1 + 10) > x_2) and ((x_1 + 10) < (x_2 + w_2)) and ((x_1 + w_1 - 10) > x_2) and ((x_1 + w_1 - 10) < (x_2 + w_2)) and ((y_1 - 10) > y_2) and ((y_1 - 10) < (y_2 + h_2)):
                        myuse.append(carID)


        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        ###4
        for carID in carTracker.keys():
            x_1 = carTracker[carID].get_position().left()
            y_1 = carTracker[carID].get_position().top()
            w_1 = carTracker[carID].get_position().width()
            h_1 = carTracker[carID].get_position().height()
            for carID_2 in carTracker.keys():
                x_2 = carTracker[carID_2].get_position().left()
                y_2 = carTracker[carID_2].get_position().top()
                w_2 = carTracker[carID_2].get_position().width()
                h_2 = carTracker[carID_2].get_position().height()


                #print(x_1, x_2, ' ', y_1, y_2, ' ', w_1, w_2, ' ', h_1, h_2)
                if (carID != carID_2):
                    #print(x_2, x_1+10, x_2+w_2, ' ', x_2, x_1+w_1-10, x_2+w_2, ' ', y_2, y_1+h_1+10, y_2+h_2, 'SEEE THIS BRO')

                    if ((x_1 - 10) > x_2) and ((x_1 - 10) < (x_2 + w_2)) and ((y_1 + 10) > y_2) and ((y_1 + 10) < (y_2 + h_2)) and ((y_1 + h_1 - 10) > y_2) and ((y_1 + h_1 - 10) < (y_2 + h_2)) :
                        myuse.append(carID)


        if myuse != []:
            #print(cars)
            for i in range(len(myuse)-1, -1, -1):
                carTracker.pop(myuse[i])
                carLocation1.pop(myuse[i])
                #currentCarID = currentCarID - 1

        #print(cars)
        myuse = []

        #print(carTracker, 'AFTERRRRRRRR')

        #b = 0
        #c = []
        
        #for carValue in carTracker.values():
            #c.append(carValue)

        #carTracker

        #for 
            

        for carID in carTracker.keys():
            trackedPosition = carTracker[carID].get_position()
					
            t_x = int(trackedPosition.left())
            t_y = int(trackedPosition.top())
            t_w = int(trackedPosition.width())
            t_h = int(trackedPosition.height())
			
            cv2.rectangle(resultImage, (t_x, t_y), (t_x + t_w, t_y + t_h), rectangleColor, 4)

            #print(t_x, ' ', t_y, ' ', t_w, ' ', t_h)
            #fig, (ax1) = plt.subplots(1)
            #ax1.imshow(resultImage)
            #plt.show()

			
			# speed estimation
            carLocation2[carID] = [t_x, t_y, t_w, t_h]
		
        end_time = time.time()
		
        if not (end_time == start_time):
            fps = 1.0/(end_time - start_time)
	    #print(fps)
		
	    #cv2.putText(resultImage, 'FPS: ' + str(int(fps)), (620, 30),cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 255), 2)


        for i in carLocation1.keys():	
            if frameCounter % 1 == 0: #it will always work
                [x1, y1, w1, h1] = carLocation1[i]
                [x2, y2, w2, h2] = carLocation2[i]
		
		# print 'previous location: ' + str(carLocation1[i]) + ', current location: ' + str(carLocation2[i])
                carLocation1[i] = [x2, y2, w2, h2]
		
		# print 'new previous location: ' + str(carLocation1[i])
                if [x1, y1, w1, h1] != [x2, y2, w2, h2]:
                    if (speed[i] == None or speed[i] == 0): #275.. 285  and y1 >= 250 and y1 <= 260.. this is to give command to the code to only estimate speed when the vehicle is between 275 and 285! 
                        speed[i] = estimateSpeed([x1, y1, w1, h1], [x2, y2, w2, h2], fps)
                        print('1st', x1, y1, w1, h1, ' ', '2nd', x2, y2, w2, h2, ' ', speed[i], ' ', 'fps:', fps)
                        #print('')

			#if y1 > 275 and y1 < 285:
                    if (speed[i] != None and (x2 > x1+5 or x2 < x1-5) and (y2 > y1+5 or y2 < y1-5)): #so that even if the driver opens his seat, the code doesn't detect speed in that
                        cv2.putText(resultImage, str(int(speed[i])) + " km/hr", (int(x1 + w1/2), int(y1-5)),cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 0, 0), 2)
                        print(speed[i], 'YOOOOOOOOOOOO')
                        print(datetime.datetime.now())
                        #if (x1 < 300):
                            #fig, (ax1) = plt.subplots(1)
                            #ax1.imshow(resultImage)
                            #plt.show()
                        

                        #mwb.save(filename)
                        
                        #Img = cv2.imread(image)
                        cropped_image[i] = image[y1:y1+h1+25, x1:x1+w1+25]
                        

                        #fig, (ax1) = plt.subplots(1)
                        #ax1.imshow(cropped_image[i])
                        #plt.show()
                        
                        LPN[i] = detection(cropped_image[i])
                        print(LPN[i])
        
                        

                        a = int(a) + int(1)
                        a = int(a)
                        c = datetime.datetime.now()
                        d = c.strftime('%d') + '/' + c.strftime('%b') + '/' + c.strftime('%Y')
                        e = c.strftime('%H') + ':' + c.strftime('%M') + ':' + c.strftime('%S')
                        

                        filename = 'C:\\Users\SHREYA\Desktop\mainFileVehicleID.xlsx'
                        mwb = load_workbook(filename)
                        sheet[counter] = mwb.active
                        max_row_sheet = sheet[counter].max_row
                        max_row_excel = int(max_row_sheet) + 1
                        sheet[counter].cell(row=max_row_excel, column=1).value = int(max_row_sheet)
                        sheet[counter].cell(row=max_row_excel, column=2).value = d
                        sheet[counter].cell(row=max_row_excel, column=3).value = e
                        sheet[counter].cell(row=max_row_excel, column=4).value = 'Cam1'
                        sheet[counter].cell(row=max_row_excel, column=5).value = speed[i]
                        sheet[counter].cell(row=max_row_excel, column=6).value = LPN[i]

                        mwb.save(filename)
                        
                        
                        #print(type(cropped_image[i]))
                        #fig, (ax1) = plt.subplots(1)
                        #ax1.imshow(cropped_image, cmap='gray')
                        #plt.show()
			#print ('CarID ' + str(i) + ': speed is ' + str("%.2f" % round(speed[i], 0)) + ' km/h.\n')

			#else:
			#	cv2.putText(resultImage, "Far Object", (int(x1 + w1/2), int(y1)),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
			#print ('CarID ' + str(i) + ' Location1: ' + str(carLocation1[i]) + ' Location2: ' + str(carLocation2[i]) + ' speed is ' + str("%.2f" % round(speed[i], 0)) + ' km/h.\n')
        cv2.imshow('result', resultImage)
	# Write the frame into the file 'output.avi'
	#out.write(resultImage)


        if cv2.waitKey(33) == 27:
            break
	
    cv2.destroyAllWindows()

if __name__ == '__main__':
    
    trackMultipleObjects()
